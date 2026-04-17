"""Synthetic test: build a DCF model with a fully populated spec
and verify zero formula errors."""
from __future__ import annotations

import sys
import re

from openpyxl import load_workbook

from mcp_servers.dcf.builder import DCFBuilder
from mcp_servers.dcf.spec import get_template, validate_spec


def make_test_spec() -> dict:
    s = get_template()
    s["company"] = {
        "name": "TestCo Industries",
        "ticker": "TESTCO.NS",
        "currency": "INR",
        "fy_end": "Mar",
    }
    s["cmp"] = 1500.0
    s["shares_diluted_mn"] = 100.0
    s["net_cash"] = -5000.0   # net debt position
    s["actuals"] = {
        "FY23": {"revenue": 50000, "ebitda": 9000, "da": 1500, "capex": 2200,
                 "nwc_change": 800, "tax_rate": 0.252,
                 "interest": 600, "avg_debt": 7500},
        "FY24": {"revenue": 58000, "ebitda": 11000, "da": 1700, "capex": 2400,
                 "nwc_change": 950, "tax_rate": 0.248,
                 "interest": 650, "avg_debt": 8200},
    }
    s["wacc"] = {
        "rf": 0.071, "erp": 0.075, "beta": 1.10,
        "kd": 0.085, "tax_rate": 0.252,
        "wt_equity": 0.85, "wt_debt": 0.15,
        "source_notes": "Rf: 10y G-Sec; ERP: Damodaran India Jul-26"
    }
    s["scenarios"] = {
        "bear": {"rev_cagr_p1": 0.07, "rev_cagr_p2": 0.04,
                  "ebitda_margin_terminal": 0.16, "tgr": 0.03,
                  "tax_rate": 0.252, "capex_pct": 0.045, "da_pct": 0.030,
                  "nwc_pct": 0.020,
                  "rationale": "Demand softness, margin compression"},
        "base": {"rev_cagr_p1": 0.12, "rev_cagr_p2": 0.07,
                  "ebitda_margin_terminal": 0.20, "tgr": 0.04,
                  "tax_rate": 0.252, "capex_pct": 0.040, "da_pct": 0.030,
                  "nwc_pct": 0.020,
                  "rationale": "Mgmt guidance: revenue 12-14% CAGR"},
        "bull": {"rev_cagr_p1": 0.18, "rev_cagr_p2": 0.10,
                  "ebitda_margin_terminal": 0.24, "tgr": 0.05,
                  "tax_rate": 0.252, "capex_pct": 0.038, "da_pct": 0.030,
                  "nwc_pct": 0.020,
                  "rationale": "Operating leverage + premiumization"},
    }
    s["projection_years"] = ["FY26", "FY27", "FY28", "FY29", "FY30",
                              "FY31", "FY32", "FY33", "FY34", "FY35"]
    s["phase1_end_idx"] = 5
    s["peers"] = [
        {"name": "Peer A", "ev_ntm_rev": 4.2, "ntm_rev_cagr": 0.14,
         "ntm_ebitda_margin": 0.22, "ev_ntm_ebitda": 19.0},
        {"name": "Peer B", "ev_ntm_rev": 3.5, "ntm_rev_cagr": 0.10,
         "ntm_ebitda_margin": 0.18, "ev_ntm_ebitda": 16.5},
        {"name": "Peer C", "ev_ntm_rev": 5.1, "ntm_rev_cagr": 0.18,
         "ntm_ebitda_margin": 0.25, "ev_ntm_ebitda": 21.0},
        {"name": "Peer D", "ev_ntm_rev": 2.8, "ntm_rev_cagr": 0.08,
         "ntm_ebitda_margin": 0.16, "ev_ntm_ebitda": 14.0},
    ]
    s["exit_multiple"] = {
        "type": "EV/EBITDA",
        "rationale": "Mature profitable cash-flow generator → EBITDA multiple",
        "bear": 12.0, "base": 17.0, "bull": 22.0,
        "metric_year": "FY30",
    }
    s["estimated_flags"] = [
        "beta: ESTIMATED — basis: 2y weekly regression on Yahoo data",
    ]
    s["key_observations"] = [
        "Base case implies modest upside vs CMP; high sensitivity to WACC",
        "Multiple-based exit yields wider range vs DCF intrinsic",
        "Reverse DCF: implied TGR exceeds Base — market pricing in stronger growth",
    ]
    return s


def main() -> int:
    spec = make_test_spec()
    errors = validate_spec(spec)
    if errors:
        print("Spec validation errors:")
        for e in errors:
            print(f"  - {e}")
        return 1
    print("✓ Spec validates")

    out = "/tmp/test_dcf.xlsx"
    DCFBuilder(spec, out).build()
    print(f"✓ Built workbook: {out}")

    # Re-open and scan for formula error strings
    wb = load_workbook(out, data_only=False)
    bad_patterns = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#NULL!|#N/A")
    issues = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                v = str(cell.value)
                if bad_patterns.search(v):
                    issues.append(f"{sheet}!{cell.coordinate}: {v}")
    if issues:
        print(f"✗ Found {len(issues)} formula errors:")
        for i in issues[:20]:
            print(f"  - {i}")
        return 1
    print("✓ No #REF!/#DIV/0!/#VALUE!/#NAME?/#NULL!/#N/A in formulas")
    print(f"✓ Sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
