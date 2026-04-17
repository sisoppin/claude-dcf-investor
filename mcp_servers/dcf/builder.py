"""DCFBuilder — build a complete multi-sheet DCF valuation xlsx.

Layout:
  Sheet 1: Cover         - summary table linked from DCF + observations + legend
  Sheet 2: Assumptions   - all inputs (blue), WACC build, scenarios, peers
  Sheet 3: DCF           - 7 sections: revenue → ebitda → fcff → wacc → val → sens → reverse
  Sheet 4: Returns_DCF   - teal: IRR matrix, exit = DCF intrinsic value
  Sheet 5: Returns_Mult  - dark red: IRR matrix on exit multiple + synthesis

All projections are formulas referencing Assumptions. No hardcoded outputs.
"""
from __future__ import annotations

from copy import copy
from typing import Any

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .styles import (
    align_center, align_indent, align_left, align_right,
    fill_darkred, fill_darkred_light, fill_header, fill_section,
    fill_teal, fill_teal_light, fill_yellow,
    FMT_INT, FMT_MULT, FMT_NUM, FMT_NUM1, FMT_PCT, FMT_PCT2, FMT_PRICE,
    font_black, font_black_bold, font_blue, font_blue_bold,
    font_green, font_green_bold, font_grey_italic, font_header,
    font_section, font_title, with_bold,
)


def _abs(sheet: str, cell: str) -> str:
    """Cross-sheet absolute reference: 'Sheet'!$B$5"""
    m = re.match(r"([A-Z]+)(\d+)$", cell)
    if not m:
        return f"'{sheet}'!{cell}"
    return f"'{sheet}'!${m.group(1)}${m.group(2)}"


def _col(c: int) -> str:
    return get_column_letter(c)


def _addr(row: int, col: int) -> str:
    return f"{_col(col)}{row}"


class DCFBuilder:
    """Build the complete DCF workbook."""

    # sheet names (kept short — referenced in formulas)
    S_COVER = "Cover"
    S_ASSUM = "Assumptions"
    S_DCF = "DCF"
    S_RDCF = "Returns_DCF"
    S_RMULT = "Returns_Multiple"

    def __init__(self, spec: dict[str, Any], output_path: str) -> None:
        self.spec = spec
        self.output_path = output_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        # named cell anchors discovered as we build
        self.A: dict[str, str] = {}      # Assumptions anchors  (e.g. A["wacc"] = "B22")
        self.D: dict[str, str] = {}      # DCF anchors

    # =================================================================
    # public
    # =================================================================
    def build(self) -> str:
        # build Assumptions first so its anchors exist
        self._build_assumptions()
        self._build_dcf()
        self._build_returns_dcf()
        self._build_returns_multiple()
        self._build_cover()  # built last (links into others), then moved to position 0
        # reorder: Cover, Assumptions, DCF, Returns_DCF, Returns_Multiple
        self.wb._sheets = [
            self.wb[self.S_COVER],
            self.wb[self.S_ASSUM],
            self.wb[self.S_DCF],
            self.wb[self.S_RDCF],
            self.wb[self.S_RMULT],
        ]
        # set print layout for every sheet: landscape, fit-to-width=1
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A3   # wider paper for DCF
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0   # auto # of pages tall
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_options.horizontalCentered = False
            ws.page_margins.left = 0.4
            ws.page_margins.right = 0.4
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            # freeze top rows for readability while scrolling
            ws.freeze_panes = "B3"
        self.wb.save(self.output_path)
        return self.output_path

    # =================================================================
    # ASSUMPTIONS
    # =================================================================
    def _build_assumptions(self) -> None:
        ws = self.wb.create_sheet(self.S_ASSUM)
        spec = self.spec
        comp = spec["company"]
        wacc = spec["wacc"]

        # column widths
        ws.column_dimensions["A"].width = 38
        for col in "BCDEFG":
            ws.column_dimensions[col].width = 16
        ws.column_dimensions["H"].width = 50

        # title
        self._title(ws, 1, f"ASSUMPTIONS — {comp.get('name', '')}", ncols=8)

        r = 3
        # ---- COMPANY ----------------------------------------------------
        self._section_bar(ws, r, "COMPANY", ncols=8); r += 1
        self._kv(ws, r, "Company name", comp.get("name", ""), is_text=True); r += 1
        self._kv(ws, r, "Ticker", comp.get("ticker", ""), is_text=True); r += 1
        self._kv(ws, r, "Currency", comp.get("currency", ""), is_text=True); r += 1
        self._kv(ws, r, "FY end", comp.get("fy_end", ""), is_text=True); r += 1
        self.A["cmp"] = self._kv(ws, r, "Current Market Price (CMP)",
                                  spec.get("cmp", 0),
                                  fmt=FMT_PRICE, highlight=True); r += 1
        self.A["shares"] = self._kv(ws, r, "Diluted shares outstanding (mn)",
                                     spec.get("shares_diluted_mn", 0),
                                     fmt=FMT_NUM1); r += 1
        self.A["net_cash"] = self._kv(ws, r, "Net cash (Cash − Debt)",
                                       spec.get("net_cash", 0),
                                       fmt=FMT_NUM); r += 1

        # ---- WACC BUILD -------------------------------------------------
        r += 1
        self._section_bar(ws, r, "WACC BUILD", ncols=8); r += 1
        # header row
        self._cell(ws, r, 1, "Component", font=font_black_bold)
        self._cell(ws, r, 2, "Value", font=font_black_bold, align=align_right)
        self._cell(ws, r, 8, "Source / Rationale", font=font_black_bold)
        r += 1

        self.A["rf"] = self._kv(ws, r, "Risk-free rate (Rf, 10y govt bond)",
                                 wacc.get("rf", 0), fmt=FMT_PCT2,
                                 source=wacc.get("source_notes", "")); r += 1
        self.A["erp"] = self._kv(ws, r, "Equity risk premium (ERP, Damodaran)",
                                  wacc.get("erp", 0), fmt=FMT_PCT2); r += 1
        self.A["beta"] = self._kv(ws, r, "Beta (2y weekly vs benchmark)",
                                   wacc.get("beta", 0), fmt='0.00'); r += 1
        # Ke = Rf + Beta*ERP  (formula, black)
        self.A["ke"] = self._formula_kv(
            ws, r, "Cost of equity (Ke = Rf + β × ERP)",
            f"={self.A['rf']}+{self.A['beta']}*{self.A['erp']}",
            fmt=FMT_PCT2); r += 1
        self.A["kd"] = self._kv(ws, r, "Cost of debt (Kd, pre-tax)",
                                 wacc.get("kd", 0), fmt=FMT_PCT2); r += 1
        self.A["tax"] = self._kv(ws, r, "Tax rate",
                                  wacc.get("tax_rate", 0), fmt=FMT_PCT2); r += 1
        self.A["kd_at"] = self._formula_kv(
            ws, r, "Kd after-tax = Kd × (1 − T)",
            f"={self.A['kd']}*(1-{self.A['tax']})", fmt=FMT_PCT2); r += 1
        self.A["wt_e"] = self._kv(ws, r, "Equity weight",
                                   wacc.get("wt_equity", 0), fmt=FMT_PCT); r += 1
        self.A["wt_d"] = self._formula_kv(
            ws, r, "Debt weight = 1 − Equity weight",
            f"=1-{self.A['wt_e']}", fmt=FMT_PCT); r += 1
        self.A["wacc"] = self._formula_kv(
            ws, r, "WACC = Ke × wₑ + Kd_AT × w_d",
            f"={self.A['ke']}*{self.A['wt_e']}+{self.A['kd_at']}*{self.A['wt_d']}",
            fmt=FMT_PCT2, highlight=True); r += 1

        # ---- SCENARIO ASSUMPTIONS --------------------------------------
        r += 1
        self._section_bar(ws, r, "SCENARIO ASSUMPTIONS", ncols=8); r += 1
        # header
        self._cell(ws, r, 1, "Driver", font=font_black_bold)
        self._cell(ws, r, 2, "Bear", font=font_header, fill=fill_header,
                   align=align_center)
        self._cell(ws, r, 3, "Base", font=font_header, fill=fill_header,
                   align=align_center)
        self._cell(ws, r, 4, "Bull", font=font_header, fill=fill_header,
                   align=align_center)
        self._cell(ws, r, 8, "Rationale", font=font_black_bold)
        r += 1

        scn_keys = [
            ("rev_cagr_p1", f"Revenue CAGR Phase 1 ({spec['projection_years'][0]}–"
                            f"{spec['projection_years'][spec['phase1_end_idx']-1]})", FMT_PCT),
            ("rev_cagr_p2", f"Revenue CAGR Phase 2 ({spec['projection_years'][spec['phase1_end_idx']]}–"
                            f"{spec['projection_years'][-1]})", FMT_PCT),
            ("ebitda_margin_terminal", "EBITDA margin (terminal year)", FMT_PCT),
            ("tax_rate", "Tax rate (steady-state)", FMT_PCT2),
            ("capex_pct", "Capex % of revenue", FMT_PCT),
            ("da_pct", "D&A % of revenue", FMT_PCT),
            ("nwc_pct", "ΔNWC % of revenue", FMT_PCT),
            ("tgr", "Terminal growth rate (TGR)", FMT_PCT2),
        ]
        for key, label, fmt in scn_keys:
            self._cell(ws, r, 1, label, font=font_black, align=align_indent)
            for ci, sc in enumerate(["bear", "base", "bull"]):
                v = spec["scenarios"][sc].get(key, 0)
                cell = self._cell(ws, r, 2 + ci, v, font=font_blue, fmt=fmt,
                                   align=align_right)
                if key == "tgr":
                    ws[cell].fill = fill_yellow
                self.A[f"{sc}_{key}"] = cell
            # rationale only on base row (use base's text)
            self._cell(ws, r, 8,
                       spec["scenarios"]["base"].get("rationale", "")
                       if key == "rev_cagr_p1" else "",
                       font=font_grey_italic)
            r += 1

        # ---- HISTORICAL ACTUALS ----------------------------------------
        r += 1
        self._section_bar(ws, r, "HISTORICAL ACTUALS", ncols=8); r += 1
        years = list(spec["actuals"].keys())
        self._cell(ws, r, 1, "Line item", font=font_black_bold)
        for i, y in enumerate(years):
            self._cell(ws, r, 2 + i, y, font=font_header, fill=fill_header,
                       align=align_center)
        self._cell(ws, r, 8, "Source", font=font_black_bold)
        r += 1

        actual_keys = [
            ("revenue", "Revenue", FMT_NUM),
            ("ebitda", "EBITDA", FMT_NUM),
            ("da", "D&A", FMT_NUM),
            ("capex", "Capex", FMT_NUM),
            ("nwc_change", "ΔNWC", FMT_NUM),
            ("tax_rate", "Effective tax rate", FMT_PCT2),
        ]
        for key, label, fmt in actual_keys:
            self._cell(ws, r, 1, label, font=font_black, align=align_indent)
            for i, y in enumerate(years):
                v = spec["actuals"][y].get(key, 0)
                cell = self._cell(ws, r, 2 + i, v, font=font_blue, fmt=fmt,
                                   align=align_right)
                self.A[f"act_{y}_{key}"] = cell
            r += 1

        # ---- PEERS ------------------------------------------------------
        r += 1
        self._section_bar(ws, r, "PEER COMPS", ncols=8); r += 1
        peer_headers = ["Peer", "EV/NTM Rev", "NTM Rev CAGR",
                        "NTM EBITDA Margin", "EV/NTM EBITDA"]
        for i, h in enumerate(peer_headers):
            self._cell(ws, r, 1 + i, h, font=font_header, fill=fill_header,
                       align=align_center if i > 0 else align_left)
        r += 1

        peer_first_row = r
        for p in spec["peers"]:
            self._cell(ws, r, 1, p["name"], font=font_blue, align=align_left)
            self._cell(ws, r, 2, p.get("ev_ntm_rev", 0), font=font_blue,
                       fmt=FMT_MULT, align=align_right)
            self._cell(ws, r, 3, p.get("ntm_rev_cagr", 0), font=font_blue,
                       fmt=FMT_PCT, align=align_right)
            self._cell(ws, r, 4, p.get("ntm_ebitda_margin", 0), font=font_blue,
                       fmt=FMT_PCT, align=align_right)
            self._cell(ws, r, 5, p.get("ev_ntm_ebitda", 0), font=font_blue,
                       fmt=FMT_MULT, align=align_right)
            r += 1
        peer_last_row = r - 1
        # mean / median rows
        self._cell(ws, r, 1, "Mean", font=font_black_bold, align=align_left)
        for ci, fmt in [(2, FMT_MULT), (3, FMT_PCT), (4, FMT_PCT), (5, FMT_MULT)]:
            self._cell(ws, r, ci,
                       f"=AVERAGE({_col(ci)}{peer_first_row}:{_col(ci)}{peer_last_row})",
                       font=font_black, fmt=fmt, align=align_right)
        r += 1
        self._cell(ws, r, 1, "Median", font=font_black_bold, align=align_left)
        for ci, fmt in [(2, FMT_MULT), (3, FMT_PCT), (4, FMT_PCT), (5, FMT_MULT)]:
            self._cell(ws, r, ci,
                       f"=MEDIAN({_col(ci)}{peer_first_row}:{_col(ci)}{peer_last_row})",
                       font=font_black, fmt=fmt, align=align_right)
        r += 1

        # ---- EXIT MULTIPLE ---------------------------------------------
        r += 1
        self._section_bar(ws, r, "EXIT MULTIPLE", ncols=8); r += 1
        em = spec["exit_multiple"]
        self._cell(ws, r, 1, "Multiple type", font=font_black, align=align_indent)
        self.A["exit_type"] = self._cell(ws, r, 2, em["type"], font=font_blue,
                                          align=align_right)
        self._cell(ws, r, 8, em.get("rationale", ""), font=font_grey_italic); r += 1
        self._cell(ws, r, 1, "Metric year", font=font_black, align=align_indent)
        self._cell(ws, r, 2, em.get("metric_year", ""), font=font_blue,
                   align=align_right); r += 1
        self._cell(ws, r, 1, "Exit multiple", font=font_black, align=align_indent)
        self._cell(ws, r, 2, "Bear", font=font_header, fill=fill_header,
                   align=align_center)
        self._cell(ws, r, 3, "Base", font=font_header, fill=fill_header,
                   align=align_center)
        self._cell(ws, r, 4, "Bull", font=font_header, fill=fill_header,
                   align=align_center); r += 1
        self._cell(ws, r, 1, "  multiple (x)", font=font_black, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            cell = self._cell(ws, r, 2 + ci, em.get(sc, 0), font=font_blue,
                               fmt=FMT_MULT, align=align_right)
            self.A[f"exit_{sc}"] = cell
        r += 1

        # ---- ESTIMATED FLAGS -------------------------------------------
        if spec.get("estimated_flags"):
            r += 1
            self._section_bar(ws, r, "ESTIMATED INPUTS (flagged)", ncols=8); r += 1
            for flag in spec["estimated_flags"]:
                self._cell(ws, r, 1, "  • " + flag, font=font_grey_italic,
                           align=align_indent); r += 1

    # =================================================================
    # DCF SHEET
    # =================================================================
    def _build_dcf(self) -> None:
        ws = self.wb.create_sheet(self.S_DCF)
        spec = self.spec
        actual_years = list(spec["actuals"].keys())   # e.g. ["FY23","FY24"]
        proj_years = spec["projection_years"]         # e.g. ["FY26"..."FY35"]
        n_act = len(actual_years)
        n_proj = len(proj_years)
        all_years = actual_years + proj_years

        # column widths
        ws.column_dimensions["A"].width = 36
        for i in range(len(all_years)):
            ws.column_dimensions[_col(2 + i)].width = 12

        # title
        self._title(ws, 1, f"DCF MODEL — {spec['company'].get('name','')}",
                    ncols=2 + len(all_years))

        # year header row (we'll repeat at major sections)
        def write_year_header(row: int) -> None:
            self._cell(ws, row, 1, "Year", font=font_black_bold)
            for i, y in enumerate(actual_years):
                self._cell(ws, row, 2 + i, f"{y}A", font=font_header,
                           fill=fill_header, align=align_center)
            for i, y in enumerate(proj_years):
                self._cell(ws, row, 2 + n_act + i, f"{y}E", font=font_header,
                           fill=fill_header, align=align_center)

        # year-number helper row (1, 2, 3, ... for projections, blank for actuals)
        def write_yearnum_row(row: int, label: str = "Year #") -> None:
            self._cell(ws, row, 1, label, font=font_grey_italic, align=align_indent)
            for i in range(n_act):
                self._cell(ws, row, 2 + i, "—", font=font_grey_italic,
                           align=align_center)
            for i in range(n_proj):
                self._cell(ws, row, 2 + n_act + i, i + 1,
                           font=font_grey_italic, fmt=FMT_INT, align=align_center)

        # ============ § 1 REVENUE BUILD ==============================
        r = 3
        self._section_bar(ws, r, "§ 1  REVENUE BUILD",
                          ncols=2 + len(all_years)); r += 1
        write_year_header(r); r += 1

        # historical revenue row (links to Assumptions)
        self._cell(ws, r, 1, "Revenue (historical)", font=font_black,
                   align=align_indent)
        for i, y in enumerate(actual_years):
            self._cell(ws, r, 2 + i,
                       f"={_abs(self.S_ASSUM, self.A[f'act_{y}_revenue'])}",
                       font=font_green, fmt=FMT_NUM, align=align_right)
        hist_rev_row = r
        r += 2

        # one row per scenario forecast revenue
        rev_rows: dict[str, int] = {}
        for sc in ["bear", "base", "bull"]:
            self._cell(ws, r, 1, f"Revenue — {sc.title()}", font=font_black_bold,
                       align=align_indent)
            for i, y in enumerate(actual_years):
                self._cell(ws, r, 2 + i,
                           f"={_addr(hist_rev_row, 2 + i)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            # FY25 (first projection) — link from last actual via CAGR p1
            last_act_col = 2 + n_act - 1
            for j in range(n_proj):
                col = 2 + n_act + j
                # which CAGR applies?
                cagr_key = ("rev_cagr_p1" if j < spec["phase1_end_idx"]
                            else "rev_cagr_p2")
                cagr_ref = _abs(self.S_ASSUM, self.A[f"{sc}_{cagr_key}"])
                prev = _addr(r, col - 1)
                self._cell(ws, r, col,
                           f"={prev}*(1+{cagr_ref})",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            rev_rows[sc] = r
            r += 1

        # growth rate %
        r += 1
        for sc in ["bear", "base", "bull"]:
            self._cell(ws, r, 1, f"  growth % — {sc.title()}",
                       font=font_grey_italic, align=align_indent)
            for j in range(n_proj):
                col = 2 + n_act + j
                cur = _addr(rev_rows[sc], col)
                prv = _addr(rev_rows[sc], col - 1)
                self._cell(ws, r, col, f"={cur}/{prv}-1",
                           font=font_grey_italic, fmt=FMT_PCT, align=align_right)
            r += 1

        # ============ § 2 EBITDA & COST WATERFALL ====================
        r += 1
        self._section_bar(ws, r, "§ 2  EBITDA & COST WATERFALL",
                          ncols=2 + len(all_years)); r += 1
        write_year_header(r); r += 1

        # historical EBITDA + margin
        self._cell(ws, r, 1, "EBITDA (historical)", font=font_black,
                   align=align_indent)
        for i, y in enumerate(actual_years):
            self._cell(ws, r, 2 + i,
                       f"={_abs(self.S_ASSUM, self.A[f'act_{y}_ebitda'])}",
                       font=font_green, fmt=FMT_NUM, align=align_right)
        hist_ebitda_row = r
        r += 1
        self._cell(ws, r, 1, "  EBITDA margin (historical)",
                   font=font_grey_italic, align=align_indent)
        for i in range(n_act):
            col = 2 + i
            self._cell(ws, r, col,
                       f"={_addr(hist_ebitda_row, col)}/{_addr(hist_rev_row, col)}",
                       font=font_grey_italic, fmt=FMT_PCT, align=align_right)
        r += 2

        ebitda_rows: dict[str, int] = {}
        margin_rows: dict[str, int] = {}
        # For each scenario: EBITDA margin ramps linearly from last historical
        # margin to terminal margin (Assumptions). EBITDA = Revenue * margin.
        for sc in ["bear", "base", "bull"]:
            # margin row first (driver)
            self._cell(ws, r, 1, f"EBITDA margin — {sc.title()}",
                       font=font_black, align=align_indent)
            # historical: same as actuals
            for i in range(n_act):
                col = 2 + i
                self._cell(ws, r, col,
                           f"={_addr(hist_ebitda_row, col)}/{_addr(hist_rev_row, col)}",
                           font=font_black, fmt=FMT_PCT, align=align_right)
            # projections: ramp linearly
            last_hist_col = 2 + n_act - 1
            term_ref = _abs(self.S_ASSUM,
                            self.A[f"{sc}_ebitda_margin_terminal"])
            for j in range(n_proj):
                col = 2 + n_act + j
                start_ref = _addr(r, last_hist_col)
                step = j + 1   # 1..n_proj
                # interpolation: start + (term - start) * step / n_proj
                self._cell(ws, r, col,
                           f"={start_ref}+({term_ref}-{start_ref})*{step}/{n_proj}",
                           font=font_black, fmt=FMT_PCT, align=align_right)
            margin_rows[sc] = r
            r += 1
            # ebitda = revenue * margin
            self._cell(ws, r, 1, f"EBITDA — {sc.title()}",
                       font=font_black_bold, align=align_indent)
            for i in range(n_act):
                col = 2 + i
                self._cell(ws, r, col,
                           f"={_addr(hist_ebitda_row, col)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            for j in range(n_proj):
                col = 2 + n_act + j
                self._cell(ws, r, col,
                           f"={_addr(rev_rows[sc], col)}*{_addr(margin_rows[sc], col)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            ebitda_rows[sc] = r
            r += 2

        # ============ § 3 FCFF BRIDGE (per scenario) =================
        r += 1
        self._section_bar(ws, r, "§ 3  FCFF BRIDGE (per scenario)",
                          ncols=2 + len(all_years)); r += 1

        fcff_rows: dict[str, int] = {}
        yearnum_rows: dict[str, int] = {}
        pv_rows: dict[str, int] = {}

        for sc in ["bear", "base", "bull"]:
            # sub-header
            self._cell(ws, r, 1, f"▸ {sc.upper()} CASE",
                       font=font_black_bold, fill=fill_section, align=align_indent)
            for c in range(2, 2 + len(all_years)):
                ws.cell(row=r, column=c).fill = fill_section
            r += 1
            write_year_header(r); r += 1
            write_yearnum_row(r); yearnum_rows[sc] = r; r += 1

            # EBITDA (link from § 2)
            self._cell(ws, r, 1, "EBITDA", font=font_black, align=align_indent)
            for i in range(len(all_years)):
                col = 2 + i
                self._cell(ws, r, col, f"={_addr(ebitda_rows[sc], col)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            ebitda_link = r; r += 1

            # − D&A
            self._cell(ws, r, 1, "  − D&A", font=font_black, align=align_indent)
            da_pct_ref = _abs(self.S_ASSUM, self.A[f"{sc}_da_pct"])
            for i in range(n_act):
                col = 2 + i
                y = actual_years[i]
                self._cell(ws, r, col,
                           f"=-{_abs(self.S_ASSUM, self.A[f'act_{y}_da'])}",
                           font=font_green, fmt=FMT_NUM, align=align_right)
            for j in range(n_proj):
                col = 2 + n_act + j
                self._cell(ws, r, col,
                           f"=-{_addr(rev_rows[sc], col)}*{da_pct_ref}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            da_row = r; r += 1

            # EBIT = EBITDA + (−D&A)
            self._cell(ws, r, 1, "EBIT", font=font_black_bold, align=align_indent)
            for i in range(len(all_years)):
                col = 2 + i
                self._cell(ws, r, col,
                           f"={_addr(ebitda_link, col)}+{_addr(da_row, col)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            ebit_row = r; r += 1

            # − Tax
            self._cell(ws, r, 1, "  − Tax", font=font_black, align=align_indent)
            tax_ref = _abs(self.S_ASSUM, self.A[f"{sc}_tax_rate"])
            for i in range(n_act):
                col = 2 + i
                y = actual_years[i]
                self._cell(ws, r, col,
                           f"=-{_addr(ebit_row, col)}*{_abs(self.S_ASSUM, self.A[f'act_{y}_tax_rate'])}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            for j in range(n_proj):
                col = 2 + n_act + j
                self._cell(ws, r, col,
                           f"=-MAX({_addr(ebit_row, col)},0)*{tax_ref}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            tax_row = r; r += 1

            # NOPAT
            self._cell(ws, r, 1, "NOPAT", font=font_black_bold, align=align_indent)
            for i in range(len(all_years)):
                col = 2 + i
                self._cell(ws, r, col,
                           f"={_addr(ebit_row, col)}+{_addr(tax_row, col)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            nopat_row = r; r += 1

            # + D&A (added back)
            self._cell(ws, r, 1, "  + D&A (add back)", font=font_black,
                       align=align_indent)
            for i in range(len(all_years)):
                col = 2 + i
                self._cell(ws, r, col, f"=-{_addr(da_row, col)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            da_back_row = r; r += 1

            # − Capex
            self._cell(ws, r, 1, "  − Capex", font=font_black, align=align_indent)
            cap_ref = _abs(self.S_ASSUM, self.A[f"{sc}_capex_pct"])
            for i in range(n_act):
                col = 2 + i
                y = actual_years[i]
                self._cell(ws, r, col,
                           f"=-{_abs(self.S_ASSUM, self.A[f'act_{y}_capex'])}",
                           font=font_green, fmt=FMT_NUM, align=align_right)
            for j in range(n_proj):
                col = 2 + n_act + j
                self._cell(ws, r, col,
                           f"=-{_addr(rev_rows[sc], col)}*{cap_ref}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            capex_row = r; r += 1

            # − ΔNWC
            self._cell(ws, r, 1, "  − ΔNWC", font=font_black, align=align_indent)
            nwc_ref = _abs(self.S_ASSUM, self.A[f"{sc}_nwc_pct"])
            for i in range(n_act):
                col = 2 + i
                y = actual_years[i]
                self._cell(ws, r, col,
                           f"=-{_abs(self.S_ASSUM, self.A[f'act_{y}_nwc_change'])}",
                           font=font_green, fmt=FMT_NUM, align=align_right)
            for j in range(n_proj):
                col = 2 + n_act + j
                self._cell(ws, r, col,
                           f"=-{_addr(rev_rows[sc], col)}*{nwc_ref}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            nwc_row = r; r += 1

            # FCFF
            self._cell(ws, r, 1, "FCFF", font=font_black_bold,
                       fill=fill_yellow, align=align_indent)
            for i in range(len(all_years)):
                col = 2 + i
                self._cell(ws, r, col,
                           f"={_addr(nopat_row, col)}+{_addr(da_back_row, col)}"
                           f"+{_addr(capex_row, col)}+{_addr(nwc_row, col)}",
                           font=font_black_bold, fmt=FMT_NUM, align=align_right,
                           fill=fill_yellow)
            fcff_rows[sc] = r; r += 1

            # Discount factor (projections only)
            self._cell(ws, r, 1, "  Discount factor", font=font_grey_italic,
                       align=align_indent)
            wacc_ref = _abs(self.S_ASSUM, self.A["wacc"])
            for i in range(n_act):
                col = 2 + i
                self._cell(ws, r, col, "—", font=font_grey_italic,
                           align=align_center)
            for j in range(n_proj):
                col = 2 + n_act + j
                yn = _addr(yearnum_rows[sc], col)
                self._cell(ws, r, col,
                           f"=1/((1+{wacc_ref})^{yn})",
                           font=font_grey_italic, fmt='0.000', align=align_right)
            disc_row = r; r += 1

            # PV of FCFF
            self._cell(ws, r, 1, "  PV of FCFF", font=font_black,
                       align=align_indent)
            for i in range(n_act):
                col = 2 + i
                self._cell(ws, r, col, "—", font=font_grey_italic,
                           align=align_center)
            for j in range(n_proj):
                col = 2 + n_act + j
                self._cell(ws, r, col,
                           f"={_addr(fcff_rows[sc], col)}*{_addr(disc_row, col)}",
                           font=font_black, fmt=FMT_NUM, align=align_right)
            pv_rows[sc] = r; r += 2

        # save FCFF rows for use elsewhere
        self.D["fcff_rows"] = fcff_rows  # type: ignore[assignment]
        self.D["pv_rows"] = pv_rows  # type: ignore[assignment]
        self.D["yearnum_rows"] = yearnum_rows  # type: ignore[assignment]
        self.D["rev_rows"] = rev_rows  # type: ignore[assignment]
        self.D["ebitda_rows"] = ebitda_rows  # type: ignore[assignment]
        self.D["n_proj"] = n_proj  # type: ignore[assignment]
        self.D["n_act"] = n_act  # type: ignore[assignment]
        self.D["proj_years"] = proj_years  # type: ignore[assignment]

        # ============ § 4 WACC SUMMARY (linked) ======================
        r += 1
        self._section_bar(ws, r, "§ 4  WACC SUMMARY (linked from Assumptions)",
                          ncols=2 + len(all_years)); r += 1
        wacc_items = [
            ("Risk-free rate (Rf)", self.A["rf"], FMT_PCT2),
            ("ERP", self.A["erp"], FMT_PCT2),
            ("Beta", self.A["beta"], '0.00'),
            ("Cost of equity (Ke)", self.A["ke"], FMT_PCT2),
            ("Cost of debt after-tax (Kd_AT)", self.A["kd_at"], FMT_PCT2),
            ("Equity weight", self.A["wt_e"], FMT_PCT),
            ("Debt weight", self.A["wt_d"], FMT_PCT),
            ("WACC", self.A["wacc"], FMT_PCT2),
        ]
        for label, anchor, fmt in wacc_items:
            self._cell(ws, r, 1, label, font=font_black, align=align_indent)
            cell = self._cell(ws, r, 2,
                              f"={_abs(self.S_ASSUM, anchor)}",
                              font=font_green, fmt=fmt, align=align_right)
            if label == "WACC":
                ws[cell].fill = fill_yellow
                ws[cell].font = with_bold(copy(font_green))
            r += 1

        # ============ § 5 VALUATION ==================================
        r += 1
        self._section_bar(ws, r, "§ 5  VALUATION (Bear / Base / Bull)",
                          ncols=2 + len(all_years)); r += 1
        # 3-column block
        self._cell(ws, r, 1, "Line item", font=font_black_bold)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci, sc.title(), font=font_header,
                       fill=fill_header, align=align_center)
        r += 1

        val_anchors: dict[str, dict[str, str]] = {sc: {} for sc in ["bear","base","bull"]}

        # Sum PV(FCFF)
        self._cell(ws, r, 1, "Sum PV(FCFFs) — explicit forecast",
                   font=font_black, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            first = _addr(pv_rows[sc], 2 + n_act)
            last = _addr(pv_rows[sc], 2 + n_act + n_proj - 1)
            val_anchors[sc]["pv_sum"] = self._cell(
                ws, r, 2 + ci, f"=SUM({first}:{last})",
                font=font_black, fmt=FMT_NUM, align=align_right)
        r += 1

        # Terminal FCFF = FCFF_N * (1+TGR)
        self._cell(ws, r, 1, "Terminal FCFF (Year N × (1+TGR))",
                   font=font_black, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            last_fcff = _addr(fcff_rows[sc], 2 + n_act + n_proj - 1)
            tgr_ref = _abs(self.S_ASSUM, self.A[f"{sc}_tgr"])
            val_anchors[sc]["term_fcff"] = self._cell(
                ws, r, 2 + ci, f"={last_fcff}*(1+{tgr_ref})",
                font=font_black, fmt=FMT_NUM, align=align_right)
        r += 1

        # Terminal Value = TFCFF / (WACC - TGR)
        self._cell(ws, r, 1, "Terminal Value (GGM)",
                   font=font_black, align=align_indent)
        wacc_ref = _abs(self.S_ASSUM, self.A["wacc"])
        for ci, sc in enumerate(["bear", "base", "bull"]):
            tgr_ref = _abs(self.S_ASSUM, self.A[f"{sc}_tgr"])
            val_anchors[sc]["tv"] = self._cell(
                ws, r, 2 + ci,
                f"={val_anchors[sc]['term_fcff']}/({wacc_ref}-{tgr_ref})",
                font=font_black, fmt=FMT_NUM, align=align_right)
        r += 1

        # PV(TV) = TV / (1+WACC)^N
        self._cell(ws, r, 1, "PV of Terminal Value",
                   font=font_black, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            val_anchors[sc]["pv_tv"] = self._cell(
                ws, r, 2 + ci,
                f"={val_anchors[sc]['tv']}/((1+{wacc_ref})^{n_proj})",
                font=font_black, fmt=FMT_NUM, align=align_right)
        r += 1

        # EV = Sum PV(FCFF) + PV(TV)
        self._cell(ws, r, 1, "Enterprise Value", font=font_black_bold,
                   align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            val_anchors[sc]["ev"] = self._cell(
                ws, r, 2 + ci,
                f"={val_anchors[sc]['pv_sum']}+{val_anchors[sc]['pv_tv']}",
                font=font_black_bold, fmt=FMT_NUM, align=align_right)
        r += 1

        # + Net Cash
        self._cell(ws, r, 1, "+ Net Cash (− Net Debt)", font=font_black,
                   align=align_indent)
        nc_ref = _abs(self.S_ASSUM, self.A["net_cash"])
        for ci, sc in enumerate(["bear", "base", "bull"]):
            val_anchors[sc]["nc"] = self._cell(
                ws, r, 2 + ci, f"={nc_ref}",
                font=font_green, fmt=FMT_NUM, align=align_right)
        r += 1

        # Equity Value
        self._cell(ws, r, 1, "Equity Value", font=font_black_bold,
                   align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            val_anchors[sc]["eq"] = self._cell(
                ws, r, 2 + ci,
                f"={val_anchors[sc]['ev']}+{val_anchors[sc]['nc']}",
                font=font_black_bold, fmt=FMT_NUM, align=align_right)
        r += 1

        # / Diluted shares
        self._cell(ws, r, 1, "÷ Diluted shares (mn)", font=font_black,
                   align=align_indent)
        sh_ref = _abs(self.S_ASSUM, self.A["shares"])
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci, f"={sh_ref}",
                       font=font_green, fmt=FMT_NUM1, align=align_right)
        r += 1

        # Implied Price (yellow)
        self._cell(ws, r, 1, "Implied Price per Share",
                   font=font_black_bold, fill=fill_yellow, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            val_anchors[sc]["price"] = self._cell(
                ws, r, 2 + ci, f"={val_anchors[sc]['eq']}/{sh_ref}",
                font=font_black_bold, fmt=FMT_PRICE, align=align_right,
                fill=fill_yellow)
        r += 1

        # CMP
        self._cell(ws, r, 1, "CMP", font=font_black, fill=fill_yellow,
                   align=align_indent)
        for ci in range(3):
            cell = self._cell(ws, r, 2 + ci,
                              f"={_abs(self.S_ASSUM, self.A['cmp'])}",
                              font=font_green, fmt=FMT_PRICE, align=align_right,
                              fill=fill_yellow)
        r += 1

        # Upside %
        self._cell(ws, r, 1, "Upside vs CMP", font=font_black_bold,
                   fill=fill_yellow, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            val_anchors[sc]["upside"] = self._cell(
                ws, r, 2 + ci,
                f"={val_anchors[sc]['price']}/{_abs(self.S_ASSUM, self.A['cmp'])}-1",
                font=font_black_bold, fmt=FMT_PCT, align=align_right,
                fill=fill_yellow)
        r += 1

        self.D["val"] = val_anchors  # type: ignore[assignment]

        # ============ § 6 SENSITIVITY 5×5 ============================
        r += 2
        self._section_bar(ws, r, "§ 6  SENSITIVITY — Implied price (Base, GGM)",
                          ncols=2 + len(all_years)); r += 1

        base_tgr_ref = _abs(self.S_ASSUM, self.A["base_tgr"])
        base_wacc_ref = wacc_ref
        base_fcff_first = _addr(self.D["fcff_rows"]["base"], 2 + n_act)        # type: ignore[index]
        base_fcff_last = _addr(self.D["fcff_rows"]["base"], 2 + n_act + n_proj - 1)  # type: ignore[index]
        base_yn_first = _addr(self.D["yearnum_rows"]["base"], 2 + n_act)       # type: ignore[index]
        base_yn_last = _addr(self.D["yearnum_rows"]["base"], 2 + n_act + n_proj - 1)  # type: ignore[index]
        nc_ref = _abs(self.S_ASSUM, self.A["net_cash"])
        sh_ref = _abs(self.S_ASSUM, self.A["shares"])

        # Sensitivity ranges: ±100bp around base, in 50bp steps (both axes)
        wacc_offsets = [-0.010, -0.005, 0.0, 0.005, 0.010]
        tgr_offsets = [-0.010, -0.005, 0.0, 0.005, 0.010]

        # Header row: corner label + TGR column headers
        self._cell(ws, r, 1, "WACC ↓  /  TGR →", font=font_black_bold,
                   align=align_center, fill=fill_section)
        for k, off in enumerate(tgr_offsets):
            self._cell(ws, r, 2 + k, f"={base_tgr_ref}+({off})",
                       font=font_black, fmt=FMT_PCT2, align=align_center,
                       fill=fill_section)
        tgr_header_row = r
        r += 1

        # Matrix: WACC row labels (col 1) + implied-price cells (cols 2-6)
        for kr, w_off in enumerate(wacc_offsets):
            self._cell(ws, r, 1, f"={base_wacc_ref}+({w_off})",
                       font=font_black, fmt=FMT_PCT2, align=align_center,
                       fill=fill_section)
            wacc_cell = f"$A${r}"
            for kc in range(len(tgr_offsets)):
                tgr_cell = f"${_col(2 + kc)}${tgr_header_row}"
                # Implied price (Base scenario, varying WACC and TGR):
                #   price = (SUMPRODUCT(FCFF / (1+WACC)^year_num)
                #            + last_FCFF*(1+TGR)/(WACC-TGR) / (1+WACC)^N
                #            + NetCash) / Shares
                formula = (
                    f"=("
                    f"SUMPRODUCT({base_fcff_first}:{base_fcff_last}"
                    f"/(1+{wacc_cell})^{base_yn_first}:{base_yn_last})"
                    f"+{base_fcff_last}*(1+{tgr_cell})"
                    f"/({wacc_cell}-{tgr_cell})/(1+{wacc_cell})^{n_proj}"
                    f"+{nc_ref}"
                    f")/{sh_ref}"
                )
                cell = self._cell(ws, r, 2 + kc, formula,
                                   font=font_black, fmt=FMT_PRICE,
                                   align=align_right)
                # highlight center cell (base WACC × base TGR)
                if kr == 2 and kc == 2:
                    ws[cell].fill = fill_yellow
                    ws[cell].font = with_bold(copy(font_black))
            r += 1
        # alias for backwards-compat with the old base_last_fcff variable used below
        base_last_fcff = base_fcff_last

        # ============ § 7 REVERSE DCF ===============================
        r += 2
        self._section_bar(ws, r, "§ 7  REVERSE DCF — what is CMP implying?",
                          ncols=2 + len(all_years)); r += 1

        cmp_ref = _abs(self.S_ASSUM, self.A["cmp"])
        sh_ref = _abs(self.S_ASSUM, self.A["shares"])
        nc_ref = _abs(self.S_ASSUM, self.A["net_cash"])
        wacc_ref = _abs(self.S_ASSUM, self.A["wacc"])

        # Implied Equity at CMP
        self._cell(ws, r, 1, "Implied Equity Value at CMP",
                   font=font_black, align=align_indent)
        ie = self._cell(ws, r, 2, f"={cmp_ref}*{sh_ref}",
                         font=font_black, fmt=FMT_NUM, align=align_right); r += 1
        # Implied EV at CMP
        self._cell(ws, r, 1, "Implied Enterprise Value at CMP",
                   font=font_black, align=align_indent)
        iev = self._cell(ws, r, 2, f"={ie}-{nc_ref}",
                          font=font_black, fmt=FMT_NUM, align=align_right); r += 1
        # DCF EV (Base)
        self._cell(ws, r, 1, "DCF EV (Base scenario)",
                   font=font_black, align=align_indent)
        dcf_ev = self._cell(ws, r, 2, f"={val_anchors['base']['ev']}",
                             font=font_green, fmt=FMT_NUM, align=align_right); r += 1
        # Gap
        self._cell(ws, r, 1, "Gap (Implied − DCF)",
                   font=font_black_bold, align=align_indent)
        self._cell(ws, r, 2, f"={iev}-{dcf_ev}",
                   font=font_black_bold, fmt=FMT_NUM, align=align_right); r += 1

        # Implied TGR (algebra: TGR = (TV*WACC - FCFF_N) / (TV + FCFF_N))
        # where TV = (Implied EV − Sum PV(FCFFs)) * (1+WACC)^N
        r += 1
        self._cell(ws, r, 1, "Implied TGR at CMP (holding Base FCFFs & WACC fixed)",
                   font=font_black_bold, align=align_indent)
        pv_sum_base = val_anchors["base"]["pv_sum"]
        last_fcff_base = base_last_fcff
        formula = (
            f"=(({iev}-{pv_sum_base})*(1+{wacc_ref})^{n_proj}*{wacc_ref}-{last_fcff_base})"
            f"/(({iev}-{pv_sum_base})*(1+{wacc_ref})^{n_proj}+{last_fcff_base})"
        )
        self._cell(ws, r, 2, formula, font=font_black_bold, fmt=FMT_PCT2,
                   fill=fill_yellow, align=align_right); r += 1

        # Implied terminal FCFF at CMP (assume Base TGR)
        self._cell(ws, r, 1, "Implied terminal FCFF at CMP (assuming Base TGR)",
                   font=font_black, align=align_indent)
        tgr_b_ref = _abs(self.S_ASSUM, self.A["base_tgr"])
        formula = (
            f"=({iev}-{pv_sum_base})*(1+{wacc_ref})^{n_proj}*"
            f"({wacc_ref}-{tgr_b_ref})/(1+{tgr_b_ref})"
        )
        self._cell(ws, r, 2, formula, font=font_black, fmt=FMT_NUM,
                   align=align_right); r += 1

        # Comparison vs explicit Base terminal FCFF
        self._cell(ws, r, 1, "  vs Base scenario terminal FCFF (explicit)",
                   font=font_grey_italic, align=align_indent)
        self._cell(ws, r, 2, f"={val_anchors['base']['term_fcff']}",
                   font=font_green, fmt=FMT_NUM, align=align_right); r += 1

    # =================================================================
    # RETURNS — DCF (teal)
    # =================================================================
    def _build_returns_dcf(self) -> None:
        ws = self.wb.create_sheet(self.S_RDCF)
        spec = self.spec
        ws.column_dimensions["A"].width = 22
        for c in range(2, 12):
            ws.column_dimensions[_col(c)].width = 13

        self._title(ws, 1,
                    f"RETURNS — DCF Intrinsic (exit = DCF intrinsic value only)",
                    fill=fill_teal, ncols=10)

        r = 3
        self._cell(ws, r, 1, "Approach: A — DCF Intrinsic",
                   font=font_black_bold, align=align_indent)
        r += 1
        self._cell(ws, r, 1,
                   "Exit prices from DCF intrinsic value only; "
                   "no market multiple applied.",
                   font=font_grey_italic, align=align_indent)
        r += 2

        # Holding periods
        holding_periods = [3, 5, 7, 10]
        # Entry prices: 7 levels around CMP (-30% to +30% in 10% steps)
        cmp_ref = _abs(self.S_ASSUM, self.A["cmp"])
        entry_offsets = [-0.30, -0.20, -0.10, 0.0, 0.10, 0.20, 0.30]

        for sc in ["bear", "base", "bull"]:
            self._cell(ws, r, 1, f"▸ {sc.upper()} CASE",
                       font=font_header, fill=fill_teal,
                       align=align_indent)
            # color the rest of the header
            for c in range(2, 2 + len(holding_periods) + 1):
                ws.cell(row=r, column=c).fill = fill_teal
                ws.cell(row=r, column=c).font = copy(font_header)
            r += 1

            # column headers
            self._cell(ws, r, 1, "Entry price ↓ / Holding period →",
                       font=font_black_bold, fill=fill_teal_light,
                       align=align_indent)
            for i, hp in enumerate(holding_periods):
                self._cell(ws, r, 2 + i, f"{hp} yr",
                           font=font_black_bold, fill=fill_teal_light,
                           align=align_center)
            self._cell(ws, r, 2 + len(holding_periods),
                       "Exit price (DCF intrinsic)",
                       font=font_black_bold, fill=fill_teal_light,
                       align=align_center)
            r += 1

            exit_ref = self.D["val"][sc]["price"]  # type: ignore[index]
            for offset in entry_offsets:
                # entry price
                entry_cell = self._cell(
                    ws, r, 1,
                    f"={cmp_ref}*(1+{offset})",
                    font=font_black, fmt=FMT_PRICE, align=align_right)
                for i, hp in enumerate(holding_periods):
                    # IRR = (exit/entry)^(1/hp) - 1
                    formula = f"=({_abs(self.S_DCF, exit_ref)}/{entry_cell})^(1/{hp})-1"
                    cell = self._cell(ws, r, 2 + i, formula,
                                       font=font_black, fmt=FMT_PCT,
                                       align=align_right)
                # exit price
                self._cell(ws, r, 2 + len(holding_periods),
                           f"={_abs(self.S_DCF, exit_ref)}",
                           font=font_green, fmt=FMT_PRICE, align=align_right)
                # highlight CMP row (offset == 0)
                if offset == 0:
                    for c in range(1, 2 + len(holding_periods) + 1):
                        ws.cell(row=r, column=c).fill = fill_yellow
                r += 1
            r += 1  # spacer

            # Max entry for target IRR
            self._cell(ws, r, 1, "Max entry for target IRR",
                       font=font_black_bold, fill=fill_teal_light,
                       align=align_indent)
            for i, hp in enumerate(holding_periods):
                self._cell(ws, r, 2 + i, f"{hp} yr",
                           font=font_black_bold, fill=fill_teal_light,
                           align=align_center)
            r += 1
            for irr in [0.10, 0.15, 0.20, 0.25, 0.30]:
                label = f"  {int(irr*100)}% IRR target"
                self._cell(ws, r, 1, label, font=font_black, align=align_indent)
                for i, hp in enumerate(holding_periods):
                    # max entry = exit / (1+irr)^hp
                    formula = f"={_abs(self.S_DCF, exit_ref)}/(1+{irr})^{hp}"
                    cell = self._cell(ws, r, 2 + i, formula,
                                       font=font_black, fmt=FMT_PRICE,
                                       align=align_right)
                if abs(irr - 0.20) < 1e-9:
                    for c in range(1, 2 + len(holding_periods)):
                        ws.cell(row=r, column=c).fill = fill_yellow
                        cur_font = ws.cell(row=r, column=c).font
                        ws.cell(row=r, column=c).font = with_bold(copy(cur_font))
                r += 1
            r += 2

    # =================================================================
    # RETURNS — Multiple (dark red)
    # =================================================================
    def _build_returns_multiple(self) -> None:
        ws = self.wb.create_sheet(self.S_RMULT)
        spec = self.spec
        em = spec["exit_multiple"]
        metric_year = em["metric_year"]

        ws.column_dimensions["A"].width = 24
        for c in range(2, 12):
            ws.column_dimensions[_col(c)].width = 14

        self._title(ws, 1, f"RETURNS — Market Multiple Exit",
                    fill=fill_darkred, ncols=10)

        r = 3
        self._cell(ws, r, 1, "Approach: B — Market Multiple Exit",
                   font=font_black_bold, align=align_indent); r += 1

        self._cell(ws, r, 1,
                   f"Multiple chosen: {em['type']} on {metric_year}. "
                   f"Rationale: {em.get('rationale','')}",
                   font=font_grey_italic, align=align_indent); r += 2

        # ---- Peer comps mini-table (link from Assumptions) ---------
        self._cell(ws, r, 1, "Peer comps (linked from Assumptions)",
                   font=font_black_bold, fill=fill_darkred_light,
                   align=align_indent)
        r += 1
        peer_headers = ["Peer", "EV/NTM Rev", "NTM Rev CAGR",
                        "NTM EBITDA Margin", "EV/NTM EBITDA"]
        for i, h in enumerate(peer_headers):
            self._cell(ws, r, 1 + i, h, font=font_header, fill=fill_darkred,
                       align=align_center if i > 0 else align_left)
        r += 1

        # We need the rows in Assumptions where peers were written.
        # Easier: just relink each peer by name + index using the order
        # in spec["peers"] starting from the known section in Assumptions.
        # We'll just link the cells we wrote: stored anchors not kept,
        # so we recompute the row offsets.
        # In _build_assumptions the peer table starts at "peer_first_row".
        # Rather than tracking that, we re-read peer values from spec
        # (they're hardcoded in Assumptions; this keeps formulas simple).
        for pi, p in enumerate(spec["peers"]):
            self._cell(ws, r, 1, p["name"], font=font_black, align=align_left)
            mult_keys = {"ev_ntm_rev", "ev_ntm_ebitda"}
            for ci, key in enumerate(["ev_ntm_rev", "ntm_rev_cagr",
                                       "ntm_ebitda_margin", "ev_ntm_ebitda"]):
                fmt = FMT_MULT if key in mult_keys else FMT_PCT
                self._cell(ws, r, 2 + ci, p.get(key, 0),
                           font=font_blue, fmt=fmt, align=align_right)
            r += 1

        r += 1
        # ---- Exit multiples (link from Assumptions) ---------------
        self._cell(ws, r, 1, "Exit multiple range",
                   font=font_black_bold, fill=fill_darkred_light,
                   align=align_indent)
        r += 1
        self._cell(ws, r, 1, "", font=font_black_bold)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci, sc.title(), font=font_header,
                       fill=fill_darkred, align=align_center)
        r += 1
        self._cell(ws, r, 1, f"Multiple ({em['type']}, {metric_year})",
                   font=font_black, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci,
                       f"={_abs(self.S_ASSUM, self.A[f'exit_{sc}'])}",
                       font=font_green, fmt=FMT_MULT, align=align_right)
        exit_mult_row = r; r += 1

        # ---- Implied price per scenario ---------------------------
        # find which projection year column corresponds to metric_year
        proj_years = self.D["proj_years"]  # type: ignore[index]
        try:
            metric_idx = proj_years.index(metric_year)
        except ValueError:
            metric_idx = len(proj_years) - 1   # fallback: last year

        n_act = self.D["n_act"]  # type: ignore[index]
        col_metric = 2 + n_act + metric_idx

        # Choose metric: revenue or EBITDA
        is_revenue = em["type"] == "EV/Revenue"
        rev_rows = self.D["rev_rows"]  # type: ignore[index]
        ebitda_rows = self.D["ebitda_rows"]  # type: ignore[index]

        self._cell(ws, r, 1, f"{('Revenue' if is_revenue else 'EBITDA')} ({metric_year})",
                   font=font_black, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            ref_row = rev_rows[sc] if is_revenue else ebitda_rows[sc]
            metric_cell = _abs(self.S_DCF, _addr(ref_row, col_metric))
            self._cell(ws, r, 2 + ci, f"={metric_cell}",
                       font=font_green, fmt=FMT_NUM, align=align_right)
        metric_row = r; r += 1

        self._cell(ws, r, 1, "Implied EV at exit", font=font_black,
                   align=align_indent)
        for ci in range(3):
            mc = _addr(metric_row, 2 + ci)
            xc = _addr(exit_mult_row, 2 + ci)
            self._cell(ws, r, 2 + ci, f"={mc}*{xc}",
                       font=font_black, fmt=FMT_NUM, align=align_right)
        ev_row = r; r += 1

        self._cell(ws, r, 1, "+ Net Cash", font=font_black, align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci,
                       f"={_abs(self.S_ASSUM, self.A['net_cash'])}",
                       font=font_green, fmt=FMT_NUM, align=align_right)
        nc_row = r; r += 1

        self._cell(ws, r, 1, "Implied Equity Value", font=font_black_bold,
                   align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci,
                       f"={_addr(ev_row, 2+ci)}+{_addr(nc_row, 2+ci)}",
                       font=font_black_bold, fmt=FMT_NUM, align=align_right)
        eq_row = r; r += 1

        self._cell(ws, r, 1, "Implied Exit Price", font=font_black_bold,
                   fill=fill_yellow, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci,
                       f"={_addr(eq_row, 2+ci)}/{_abs(self.S_ASSUM, self.A['shares'])}",
                       font=font_black_bold, fmt=FMT_PRICE,
                       fill=fill_yellow, align=align_right)
        exit_price_row = r; r += 2

        # ---- IRR sensitivity: entry × multiple (Base) -------------
        self._cell(ws, r, 1, "IRR sensitivity (Base scenario)",
                   font=font_black_bold, fill=fill_darkred_light,
                   align=align_indent); r += 1
        self._cell(ws, r, 1, "Entry ↓ / Exit multiple →",
                   font=font_black_bold, align=align_indent)

        # multiple range: ±30% around base in 4 steps
        base_mult_ref = _addr(exit_mult_row, 3)  # base column
        for k in range(5):
            offset = -0.30 + k * 0.15
            self._cell(ws, r, 2 + k, f"={base_mult_ref}*(1+{offset})",
                       font=font_black, fmt=FMT_MULT,
                       fill=fill_darkred_light, align=align_center)
        r += 1

        # Holding period (assume metric_year offset from now ~ 5y)
        # Use the metric_year index + 1 as a default holding period
        hp = metric_idx + 1
        cmp_ref = _abs(self.S_ASSUM, self.A["cmp"])
        nc_ref = _abs(self.S_ASSUM, self.A["net_cash"])
        sh_ref = _abs(self.S_ASSUM, self.A["shares"])
        metric_base_cell = _addr(metric_row, 3)  # base metric

        self._cell(ws, r, 1, f"(holding period assumed: {hp}y)",
                   font=font_grey_italic, align=align_indent)
        r += 1
        # mult_header_row is 2 rows above r (the multiple-axis header we wrote earlier)
        mult_header_row = r - 2
        entry_offsets = [-0.30, -0.20, -0.10, 0.0, 0.10, 0.20, 0.30]
        for offset in entry_offsets:
            entry = self._cell(
                ws, r, 1, f"={cmp_ref}*(1+{offset})",
                font=font_black, fmt=FMT_PRICE, align=align_right)
            for k in range(5):
                mult_cell = f"${_col(2 + k)}${mult_header_row}"
                # Implied exit price = (Base_metric × mult + NetCash) / Shares
                exit_price = f"({metric_base_cell}*{mult_cell}+{nc_ref})/{sh_ref}"
                # IRR = (exit/entry)^(1/hp) - 1
                self._cell(ws, r, 2 + k,
                           f"=({exit_price}/{entry})^(1/{hp})-1",
                           font=font_black, fmt=FMT_PCT, align=align_right)
            # highlight the CMP row (offset == 0)
            if abs(offset) < 1e-9:
                for c in range(1, 7):
                    ws.cell(row=r, column=c).fill = fill_yellow
            r += 1
        r += 1

        # Max entry for target IRRs
        self._cell(ws, r, 1, "Max entry for target IRR (Base multiple)",
                   font=font_black_bold, fill=fill_darkred_light,
                   align=align_indent); r += 1
        # exit price at base multiple = exit_price_row, base column
        base_exit = _addr(exit_price_row, 3)
        self._cell(ws, r, 1, "Target IRR →", font=font_black_bold,
                   align=align_indent)
        for ci, irr in enumerate([0.10, 0.15, 0.20, 0.25, 0.30]):
            self._cell(ws, r, 2 + ci, f"{int(irr*100)}%",
                       font=font_black_bold, align=align_center)
        r += 1
        self._cell(ws, r, 1, f"Max entry ({hp}y holding)",
                   font=font_black, align=align_indent)
        for ci, irr in enumerate([0.10, 0.15, 0.20, 0.25, 0.30]):
            cell = self._cell(ws, r, 2 + ci,
                               f"={base_exit}/(1+{irr})^{hp}",
                               font=font_black, fmt=FMT_PRICE,
                               align=align_right)
            if abs(irr - 0.20) < 1e-9:
                ws[cell].fill = fill_yellow
                ws[cell].font = with_bold(copy(font_black))
        r += 2

        # ---- Synthesis: A vs B ---------------------------------------
        self._cell(ws, r, 1, "SYNTHESIS — Approach A (DCF) vs B (Multiple)",
                   font=font_black_bold, fill=fill_darkred_light,
                   align=align_indent); r += 1
        headers = ["", "Bear", "Base", "Bull"]
        for ci, h in enumerate(headers):
            self._cell(ws, r, 1 + ci, h, font=font_header, fill=fill_darkred,
                       align=align_center if ci > 0 else align_left)
        r += 1

        # A: DCF implied price
        self._cell(ws, r, 1, "A) DCF intrinsic price",
                   font=font_black, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci,
                       f"={_abs(self.S_DCF, self.D['val'][sc]['price'])}",  # type: ignore[index]
                       font=font_green, fmt=FMT_PRICE, align=align_right)
        a_row = r; r += 1

        # B: Multiple implied price (from exit_price_row above)
        self._cell(ws, r, 1, "B) Multiple-based exit price",
                   font=font_black, align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci, f"={_addr(exit_price_row, 2+ci)}",
                       font=font_black, fmt=FMT_PRICE, align=align_right)
        b_row = r; r += 1

        self._cell(ws, r, 1, "Upside vs CMP — A",
                   font=font_black, align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci,
                       f"={_addr(a_row, 2+ci)}/{cmp_ref}-1",
                       font=font_black, fmt=FMT_PCT, align=align_right)
        r += 1
        self._cell(ws, r, 1, "Upside vs CMP — B",
                   font=font_black, align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci,
                       f"={_addr(b_row, 2+ci)}/{cmp_ref}-1",
                       font=font_black, fmt=FMT_PCT, align=align_right)
        r += 1

        # IRR at CMP for both, base scenario, hp = metric_idx+1
        self._cell(ws, r, 1, f"IRR at CMP (assumes {hp}y holding) — A",
                   font=font_black, align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci,
                       f"=({_addr(a_row, 2+ci)}/{cmp_ref})^(1/{hp})-1",
                       font=font_black, fmt=FMT_PCT, align=align_right)
        r += 1
        self._cell(ws, r, 1, f"IRR at CMP (assumes {hp}y holding) — B",
                   font=font_black, align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci,
                       f"=({_addr(b_row, 2+ci)}/{cmp_ref})^(1/{hp})-1",
                       font=font_black, fmt=FMT_PCT, align=align_right)
        r += 1

    # =================================================================
    # COVER
    # =================================================================
    def _build_cover(self) -> None:
        ws = self.wb.create_sheet(self.S_COVER)
        spec = self.spec
        ws.column_dimensions["A"].width = 36
        for c in "BCDE":
            ws.column_dimensions[c].width = 18

        self._title(ws, 1, f"{spec['company'].get('name', '')} — DCF Valuation",
                    ncols=5)

        r = 3
        # Color legend
        self._cell(ws, r, 1, "Color legend", font=font_black_bold,
                   align=align_indent); r += 1
        legend = [
            ("Blue text", "hardcoded input (Assumptions only)", font_blue),
            ("Black text", "formula", font_black),
            ("Green text", "cross-sheet link", font_green),
            ("Yellow background", "key cells / CMP / 20% IRR row", None),
        ]
        for label, desc, fnt in legend:
            cell = self._cell(ws, r, 1, "  " + label,
                               font=fnt or font_black, align=align_indent)
            if "Yellow" in label:
                ws[cell].fill = fill_yellow
            self._cell(ws, r, 2, desc, font=font_grey_italic, align=align_left)
            r += 1
        r += 1

        # Bear/Base/Bull summary
        self._section_bar(ws, r, "VALUATION SUMMARY", ncols=5); r += 1
        self._cell(ws, r, 1, "Metric", font=font_black_bold)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci, sc.title(), font=font_header,
                       fill=fill_header, align=align_center)
        r += 1

        # Last forecast year revenue
        n_act = self.D["n_act"]  # type: ignore[index]
        n_proj = self.D["n_proj"]  # type: ignore[index]
        last_col = 2 + n_act + n_proj - 1
        rev_rows = self.D["rev_rows"]  # type: ignore[index]
        ebitda_rows = self.D["ebitda_rows"]  # type: ignore[index]
        fcff_rows = self.D["fcff_rows"]  # type: ignore[index]
        proj_years = self.D["proj_years"]  # type: ignore[index]

        terminal_year = proj_years[-1]
        rows_spec = [
            (f"Revenue ({terminal_year})", rev_rows, FMT_NUM),
            (f"EBITDA ({terminal_year})", ebitda_rows, FMT_NUM),
            (f"FCFF ({terminal_year})", fcff_rows, FMT_NUM),
        ]
        for label, rmap, fmt in rows_spec:
            self._cell(ws, r, 1, label, font=font_black, align=align_indent)
            for ci, sc in enumerate(["bear", "base", "bull"]):
                self._cell(ws, r, 2 + ci,
                           f"={_abs(self.S_DCF, _addr(rmap[sc], last_col))}",
                           font=font_green, fmt=fmt, align=align_right)
            r += 1

        val = self.D["val"]  # type: ignore[index]
        for label, key, fmt in [
            ("Sum PV(FCFFs)", "pv_sum", FMT_NUM),
            ("Terminal Value", "tv", FMT_NUM),
            ("Enterprise Value", "ev", FMT_NUM),
            ("Equity Value", "eq", FMT_NUM),
        ]:
            self._cell(ws, r, 1, label, font=font_black, align=align_indent)
            for ci, sc in enumerate(["bear", "base", "bull"]):
                self._cell(ws, r, 2 + ci,
                           f"={_abs(self.S_DCF, val[sc][key])}",
                           font=font_green, fmt=fmt, align=align_right)
            r += 1

        # Implied price (yellow)
        self._cell(ws, r, 1, "Implied Price per Share",
                   font=font_black_bold, fill=fill_yellow,
                   align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci,
                       f"={_abs(self.S_DCF, val[sc]['price'])}",
                       font=font_green_bold, fmt=FMT_PRICE,
                       fill=fill_yellow, align=align_right)
        r += 1

        # CMP (yellow)
        self._cell(ws, r, 1, "CMP", font=font_black, fill=fill_yellow,
                   align=align_indent)
        for ci in range(3):
            self._cell(ws, r, 2 + ci,
                       f"={_abs(self.S_ASSUM, self.A['cmp'])}",
                       font=font_green, fmt=FMT_PRICE,
                       fill=fill_yellow, align=align_right)
        r += 1

        # Upside %
        self._cell(ws, r, 1, "Upside vs CMP", font=font_black_bold,
                   fill=fill_yellow, align=align_indent)
        for ci, sc in enumerate(["bear", "base", "bull"]):
            self._cell(ws, r, 2 + ci,
                       f"={_abs(self.S_DCF, val[sc]['upside'])}",
                       font=font_green_bold, fmt=FMT_PCT,
                       fill=fill_yellow, align=align_right)
        r += 2

        # Key observations
        self._section_bar(ws, r, "KEY OBSERVATIONS", ncols=5); r += 1
        for obs in (spec.get("key_observations") or
                    ["(no observations provided)"]):
            self._cell(ws, r, 1, "  • " + obs,
                       font=font_black, align=align_indent)
            r += 1

        # Estimated flags
        if spec.get("estimated_flags"):
            r += 1
            self._section_bar(ws, r, "ESTIMATED INPUTS (flagged)", ncols=5); r += 1
            for flag in spec["estimated_flags"]:
                self._cell(ws, r, 1, "  • " + flag, font=font_grey_italic,
                           align=align_indent)
                r += 1

    # =================================================================
    # primitives
    # =================================================================
    def _cell(self, ws: Worksheet, row: int, col: int, value: Any, *,
              font=None, fill=None, fmt: str | None = None,
              align=None) -> str:
        c = ws.cell(row=row, column=col, value=value)
        if font is not None:
            c.font = copy(font)
        if fill is not None:
            c.fill = fill
        if fmt is not None:
            c.number_format = fmt
        if align is not None:
            c.alignment = align
        return _addr(row, col)

    def _title(self, ws: Worksheet, row: int, text: str, *,
               fill=None, ncols: int = 5) -> None:
        used_fill = fill or fill_header
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = used_fill
        c0 = ws.cell(row=row, column=1, value=text)
        c0.font = copy(font_title)
        c0.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def _section_bar(self, ws: Worksheet, row: int, text: str, *,
                     ncols: int = 5) -> None:
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = fill_section
        c0 = ws.cell(row=row, column=1, value=text)
        c0.font = copy(font_section)
        c0.alignment = align_indent

    def _kv(self, ws: Worksheet, row: int, label: str, value: Any, *,
            fmt: str | None = None, is_text: bool = False,
            highlight: bool = False, source: str = "") -> str:
        self._cell(ws, row, 1, label, font=font_black, align=align_indent)
        cell = self._cell(ws, row, 2, value, font=font_blue, fmt=fmt,
                          align=align_right if not is_text else align_left)
        if highlight:
            ws[cell].fill = fill_yellow
            ws[cell].font = with_bold(copy(font_blue))
        if source:
            self._cell(ws, row, 8, source, font=font_grey_italic,
                       align=align_left)
        return cell

    def _formula_kv(self, ws: Worksheet, row: int, label: str, formula: str, *,
                    fmt: str | None = None, highlight: bool = False) -> str:
        self._cell(ws, row, 1, label, font=font_black, align=align_indent)
        cell = self._cell(ws, row, 2, formula, font=font_black, fmt=fmt,
                          align=align_right)
        if highlight:
            ws[cell].fill = fill_yellow
            ws[cell].font = with_bold(copy(font_black))
        return cell
