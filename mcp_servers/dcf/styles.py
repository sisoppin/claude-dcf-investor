"""Color palette, fonts, fills, and number formats for the DCF workbook."""
from __future__ import annotations

from copy import copy

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---- colors ---------------------------------------------------------
BLUE = "0070C0"            # hardcoded inputs
GREEN = "00B050"           # cross-sheet links
BLACK = "000000"           # formulas
WHITE = "FFFFFF"
GREY = "595959"

YELLOW_BG = "FFF2CC"       # key cells, CMP row, 20% IRR row
HEADER_BG = "1F4E78"       # main headers (dark blue)
SECTION_BG = "D9E1F2"      # section bars (light blue)

TEAL_BG = "0E6E6E"         # Returns DCF (teal scheme)
TEAL_LIGHT = "B2DFDB"
DARKRED_BG = "8B0000"      # Returns Multiple (dark red scheme)
DARKRED_LIGHT = "F4CCCC"

FONT = "Calibri"

# ---- fonts ----------------------------------------------------------
font_blue = Font(name=FONT, color=BLUE)
font_green = Font(name=FONT, color=GREEN)
font_black = Font(name=FONT, color=BLACK)
font_grey_italic = Font(name=FONT, color=GREY, italic=True, size=9)

font_header = Font(name=FONT, color=WHITE, bold=True, size=11)
font_section = Font(name=FONT, color=BLACK, bold=True, size=11)
font_title = Font(name=FONT, color=WHITE, bold=True, size=14)

font_blue_bold = Font(name=FONT, color=BLUE, bold=True)
font_green_bold = Font(name=FONT, color=GREEN, bold=True)
font_black_bold = Font(name=FONT, color=BLACK, bold=True)

# ---- fills ----------------------------------------------------------
fill_yellow = PatternFill("solid", fgColor=YELLOW_BG)
fill_header = PatternFill("solid", fgColor=HEADER_BG)
fill_section = PatternFill("solid", fgColor=SECTION_BG)
fill_teal = PatternFill("solid", fgColor=TEAL_BG)
fill_teal_light = PatternFill("solid", fgColor=TEAL_LIGHT)
fill_darkred = PatternFill("solid", fgColor=DARKRED_BG)
fill_darkred_light = PatternFill("solid", fgColor=DARKRED_LIGHT)

# ---- borders --------------------------------------------------------
_thin = Side(style="thin", color="BFBFBF")
border_box = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
border_top = Border(top=_thin)
border_bottom = Border(bottom=_thin)
border_top_thick = Border(top=Side(style="medium", color="000000"))

# ---- alignments -----------------------------------------------------
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_indent = Alignment(horizontal="left", vertical="center", indent=1)

# ---- number formats -------------------------------------------------
FMT_NUM = '#,##0;(#,##0);"–"'
FMT_NUM1 = '#,##0.0;(#,##0.0);"–"'
FMT_PCT = '0.0%;(0.0%);"–"'
FMT_PCT2 = '0.00%;(0.00%);"–"'
FMT_PRICE = '#,##0.00;(#,##0.00);"–"'
FMT_MULT = '0.0"x";(0.0"x");"–"'
FMT_INT = '0'


def with_bold(font: Font) -> Font:
    f = copy(font)
    f.bold = True
    return f
