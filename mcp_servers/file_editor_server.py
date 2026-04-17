"""Local MCP server: read/write/edit json, yaml, csv, xlsx files.

All paths are sandboxed under WORKSPACE_PATH (default ./workspace).

Tools:
  - list_files(subdir)
  - read_file(path)                       (auto-detects format)
  - write_file(path, content)             (raw text/json string write)
  - edit_json(path, json_path, value)
  - edit_yaml(path, yaml_path, value)
  - read_csv(path, max_rows)
  - append_csv(path, row)                 (row is list[str] or dict)
  - read_xlsx(path, sheet, max_rows)
  - write_xlsx(path, sheet, rows)         (rows = list[dict] or list[list])
  - delete_file(path)
"""
from __future__ import annotations

import csv
import io
import json
import os
from pathlib import Path
from typing import Any

import yaml
from jsonpath_ng import parse as jsonpath_parse
from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook

mcp = FastMCP("file_editor")

WORKSPACE = Path(os.getenv("WORKSPACE_PATH", "./workspace")).expanduser().resolve()
WORKSPACE.mkdir(parents=True, exist_ok=True)


# ---------- helpers --------------------------------------------------------


def _resolve(path: str) -> Path:
    """Resolve a user-supplied path inside the workspace; reject escapes."""
    if not path:
        raise ValueError("path is required")
    p = (WORKSPACE / path).resolve()
    if WORKSPACE not in p.parents and p != WORKSPACE:
        raise ValueError(f"path '{path}' escapes workspace")
    return p


def _ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)


def _format_table(rows: list[list[Any]], header: list[str] | None = None) -> str:
    out = io.StringIO()
    writer = csv.writer(out)
    if header:
        writer.writerow(header)
    for row in rows:
        writer.writerow(row)
    return out.getvalue().rstrip()


# ---------- tools ----------------------------------------------------------


@mcp.tool()
def list_files(subdir: str = "") -> str:
    """List files (recursively) under the workspace, optionally scoped to a subdir."""
    base = _resolve(subdir) if subdir else WORKSPACE
    if not base.exists():
        return f"(no such directory: {subdir})"
    entries: list[str] = []
    for p in sorted(base.rglob("*")):
        if p.is_file():
            entries.append(str(p.relative_to(WORKSPACE)))
    return "\n".join(entries) if entries else "(workspace is empty)"


@mcp.tool()
def read_file(path: str) -> str:
    """Read a file. Auto-detects json/yaml/csv/xlsx and pretty-prints the content."""
    p = _resolve(path)
    if not p.exists():
        return f"ERROR: file not found: {path}"

    suffix = p.suffix.lower()
    try:
        if suffix == ".json":
            return json.dumps(json.loads(p.read_text(encoding="utf-8")), indent=2)
        if suffix in {".yaml", ".yml"}:
            data = yaml.safe_load(p.read_text(encoding="utf-8"))
            return yaml.safe_dump(data, sort_keys=False, allow_unicode=True)
        if suffix == ".csv":
            with p.open(newline="", encoding="utf-8") as f:
                return f.read()
        if suffix == ".xlsx":
            return read_xlsx(path)
        # text fallback
        return p.read_text(encoding="utf-8")
    except Exception as e:  # noqa: BLE001
        return f"ERROR reading {path}: {e}"


@mcp.tool()
def write_file(path: str, content: str) -> str:
    """Write raw text content to a file (creates parent dirs).

    For structured edits to existing files, prefer edit_json / edit_yaml /
    append_csv / write_xlsx.
    """
    p = _resolve(path)
    _ensure_parent(p)
    try:
        p.write_text(content, encoding="utf-8")
        return f"wrote {len(content)} chars to {path}"
    except Exception as e:  # noqa: BLE001
        return f"ERROR writing {path}: {e}"


@mcp.tool()
def edit_json(path: str, json_path: str, value: Any) -> str:
    """Set a value at a JSONPath inside a JSON file.

    Example: json_path='$.users[0].name', value='Alice'
    Creates the file with {} if it doesn't exist.
    """
    p = _resolve(path)
    _ensure_parent(p)
    try:
        data = json.loads(p.read_text(encoding="utf-8")) if p.exists() else {}
        expr = jsonpath_parse(json_path)
        matches = expr.find(data)
        if not matches:
            # update_or_create handles creating missing keys when possible
            expr.update_or_create(data, value)
        else:
            expr.update(data, value)
        p.write_text(json.dumps(data, indent=2), encoding="utf-8")
        return f"updated {json_path} in {path}"
    except Exception as e:  # noqa: BLE001
        return f"ERROR editing {path}: {e}"


@mcp.tool()
def edit_yaml(path: str, yaml_path: str, value: Any) -> str:
    """Set a value at a dotted path inside a YAML file.

    Example: yaml_path='database.host', value='localhost'
    Use 'list[0]' style for list indices.
    Creates the file if it doesn't exist.
    """
    p = _resolve(path)
    _ensure_parent(p)
    try:
        data = yaml.safe_load(p.read_text(encoding="utf-8")) if p.exists() else {}
        if data is None:
            data = {}
        _set_dotted(data, yaml_path, value)
        p.write_text(
            yaml.safe_dump(data, sort_keys=False, allow_unicode=True), encoding="utf-8"
        )
        return f"updated {yaml_path} in {path}"
    except Exception as e:  # noqa: BLE001
        return f"ERROR editing {path}: {e}"


def _set_dotted(obj: Any, path: str, value: Any) -> None:
    """Walk a dotted path with optional [i] indices and set the leaf."""
    import re

    tokens = re.findall(r"[^.\[\]]+|\[\d+\]", path)
    cur = obj
    for i, tok in enumerate(tokens):
        last = i == len(tokens) - 1
        is_index = tok.startswith("[") and tok.endswith("]")
        key: Any = int(tok[1:-1]) if is_index else tok

        if last:
            if is_index:
                while len(cur) <= key:
                    cur.append(None)
                cur[key] = value
            else:
                cur[key] = value
            return

        next_tok = tokens[i + 1]
        next_is_index = next_tok.startswith("[") and next_tok.endswith("]")
        default: Any = [] if next_is_index else {}

        if is_index:
            while len(cur) <= key:
                cur.append(default)
            if cur[key] is None:
                cur[key] = default
            cur = cur[key]
        else:
            if key not in cur or cur[key] is None:
                cur[key] = default
            cur = cur[key]


@mcp.tool()
def read_csv(path: str, max_rows: int = 100) -> str:
    """Read a CSV and return up to max_rows rows."""
    p = _resolve(path)
    if not p.exists():
        return f"ERROR: file not found: {path}"
    try:
        with p.open(newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = []
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(row)
        return _format_table(rows)
    except Exception as e:  # noqa: BLE001
        return f"ERROR reading {path}: {e}"


@mcp.tool()
def append_csv(path: str, row: Any) -> str:
    """Append a row to a CSV file. Row may be a list of values or a dict.

    If the file doesn't exist and row is a dict, a header is written first.
    """
    p = _resolve(path)
    _ensure_parent(p)
    try:
        is_dict = isinstance(row, dict)
        new_file = not p.exists()
        with p.open("a", newline="", encoding="utf-8") as f:
            if is_dict:
                writer = csv.DictWriter(f, fieldnames=list(row.keys()))
                if new_file:
                    writer.writeheader()
                writer.writerow(row)
            else:
                csv.writer(f).writerow(row)
        return f"appended row to {path}"
    except Exception as e:  # noqa: BLE001
        return f"ERROR appending to {path}: {e}"


@mcp.tool()
def read_xlsx(path: str, sheet: str = "", max_rows: int = 100) -> str:
    """Read an XLSX sheet and return it as CSV-like text."""
    p = _resolve(path)
    if not p.exists():
        return f"ERROR: file not found: {path}"
    try:
        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(["" if v is None else v for v in row])
        sheet_name = sheet or wb.sheetnames[0]
        return f"# sheet: {sheet_name}\n" + _format_table(rows)
    except Exception as e:  # noqa: BLE001
        return f"ERROR reading {path}: {e}"


@mcp.tool()
def write_xlsx(path: str, sheet: str, rows: list) -> str:
    """Write rows to an XLSX sheet (overwrites the sheet).

    rows: either list[list] (raw rows) or list[dict] (keys become header).
    Other sheets in an existing workbook are preserved.
    """
    p = _resolve(path)
    _ensure_parent(p)
    try:
        if p.exists():
            wb = load_workbook(p)
            if sheet in wb.sheetnames:
                del wb[sheet]
        else:
            wb = Workbook()
            # remove the default empty sheet if present and unused
            default = wb.active
            if default is not None and default.title == "Sheet" and default.max_row == 1:
                wb.remove(default)
        ws = wb.create_sheet(title=sheet)

        if rows and isinstance(rows[0], dict):
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h, "") for h in headers])
        else:
            for r in rows:
                ws.append(list(r))
        wb.save(p)
        return f"wrote {len(rows)} rows to {path} [sheet: {sheet}]"
    except Exception as e:  # noqa: BLE001
        return f"ERROR writing {path}: {e}"


@mcp.tool()
def delete_file(path: str) -> str:
    """Delete a file from the workspace."""
    p = _resolve(path)
    if not p.exists():
        return f"(already absent: {path})"
    try:
        p.unlink()
        return f"deleted {path}"
    except Exception as e:  # noqa: BLE001
        return f"ERROR deleting {path}: {e}"


if __name__ == "__main__":
    mcp.run()
